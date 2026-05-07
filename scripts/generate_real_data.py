from __future__ import annotations

import json
import re
from pathlib import Path
from urllib.parse import quote

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PENDING = "\u5f85\u8865\u5145"
RESOURCE_XLSX = ROOT / "\u5317\u4eac\u6587\u5316\u8d44\u6e90\u5bfc\u822a_\u53bb\u9664\u8857\u9053\u7efc\u5408\u6587\u5316\u4e2d\u5fc3_\u53bb\u91cd\u6574\u7406.xlsx"
EVENT_XLSX = ROOT / "\u5317\u4eac\u5546\u4e1a\u6f14\u51fa\u4fe1\u606f_2026-05-04\u81f305-10_\u7f51\u7ad9\u6570\u636e\u6574\u7406.xlsx"


def text(value) -> str:
    return "" if value is None else str(value).strip()


def dump(value) -> str:
    return json.dumps(value, ensure_ascii=True)


def baidu(name: str) -> str:
    return f"https://map.baidu.com/search/{quote(name or '')}"


def amap(name: str) -> str:
    return f"https://www.amap.com/search?query={quote(name or '')}"


def split_tags(value: str) -> list[str]:
    return [item.strip() for item in re.split(r"[;\uff1b\u3001/\uff0c,]+", text(value)) if item and item.strip()]


RESOURCE_TYPE_MAP = {
    "\u516c\u5171\u56fe\u4e66\u9986": "\u56fe\u4e66\u9986",
    "\u6587\u5316\u9986": "\u6587\u5316\u9986",
    "\u6f14\u51fa\u573a\u6240": "\u5267\u573a/\u6f14\u51fa",
    "\u827a\u672f\u8868\u6f14\u56e2\u4f53": "\u5267\u573a/\u6f14\u51fa",
    "\u6587\u5316\u573a\u6240": "\u5176\u4ed6\u6587\u5316\u7a7a\u95f4",
}


def map_resource_type(value: str) -> str:
    return RESOURCE_TYPE_MAP.get(text(value), "\u5176\u4ed6\u6587\u5316\u7a7a\u95f4")


def unique(values: list[str]) -> list[str]:
    seen = set()
    result = []
    for value in values:
        if value and value not in seen:
            seen.add(value)
            result.append(value)
    return result


def audiences_from(tags: list[str], resource_type: str) -> list[str]:
    joined = " ".join(tags + [resource_type])
    result = []
    if any(key in joined for key in ["\u5b66\u751f", "\u81ea\u4e60", "\u9605\u8bfb", "\u8bb2\u5ea7"]):
        result.append("\u5b66\u751f")
    if any(key in joined for key in ["\u591c\u95f4", "\u6f14\u51fa", "\u5c55\u89c8", "\u9605\u8bfb"]):
        result.append("\u4e0a\u73ed\u65cf")
    if any(key in joined for key in ["\u620f\u66f2", "\u66f2\u827a", "\u4e66\u753b", "\u516c\u5171\u6587\u5316", "\u6587\u5316\u9986"]):
        result.append("\u8001\u5e74\u4eba")
    if any(key in joined for key in ["\u4eb2\u5b50", "\u513f\u7ae5", "\u5c55\u89c8", "\u56fe\u4e66\u9986"]):
        result.append("\u4eb2\u5b50\u5bb6\u5ead")
    if "\u65e0\u969c\u788d" in joined:
        result.append("\u6b8b\u969c\u4eba\u58eb")
    if any(key in joined for key in ["\u666f\u533a", "\u535a\u7269\u9986", "\u6587\u5316", "\u6f14\u51fa", "\u5c55\u89c8"]):
        result.append("\u65b0\u5317\u4eac\u4eba/\u6e38\u5ba2")
    return result or ["\u65b0\u5317\u4eac\u4eba/\u6e38\u5ba2"]


def build_resources() -> list[dict]:
    wb = load_workbook(RESOURCE_XLSX, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if not any(raw):
            continue
        (
            rid,
            name,
            normalized,
            level_type,
            activity_type,
            district,
            address,
            phone,
            open_time,
            tags_raw,
            official_url,
            activity_url,
            ticket_url,
            map_url,
            source,
        ) = [text(item) for item in raw[:15]]
        display_name = normalized or name
        resource_type = map_resource_type(level_type)
        tags = unique(split_tags(activity_type) + split_tags(tags_raw) + [level_type, source])
        source_text = source or PENDING
        activity_text = activity_type or PENDING
        resource = {
            "id": rid or f"resource-{len(rows) + 1}",
            "name": display_name,
            "type": resource_type,
            "district": district or "\u5f85\u8865\u5145",
            "address": address or "\u5f85\u8865\u5145",
            "phone": phone or "\u5f85\u8865\u5145",
            "openTime": open_time or "\u4ee5\u5b98\u65b9\u4e3a\u51c6",
            "tags": tags,
            "audiences": audiences_from(tags, resource_type),
            "intro": f"{display_name}\uff0c\u6765\u6e90\u7c7b\u522b\uff1a{source_text}\u3002\u8d44\u6e90/\u6d3b\u52a8\u7c7b\u578b\uff1a{activity_text}\u3002\u5f00\u653e\u3001\u9884\u7ea6\u3001\u6d3b\u52a8\u548c\u8def\u7ebf\u4fe1\u606f\u8bf7\u4ee5\u5b98\u65b9\u9875\u9762\u6216\u73b0\u573a\u5b9e\u9645\u60c5\u51b5\u4e3a\u51c6\u3002",
            "officialUrl": official_url,
            "activityUrl": activity_url,
            "ticketUrl": ticket_url,
            "mapUrl": map_url or baidu(display_name),
            "amapUrl": amap(f"{display_name} {address}".strip()),
            "isFree": any("\u514d\u8d39" in tag or "\u4f4e\u4ef7" in tag for tag in tags),
            "source": source or "\u5317\u4eac\u6587\u5316\u8d44\u6e90\u5bfc\u822a\u6574\u7406\u8868",
            "verifyStatus": "\u5b98\u65b9\u6765\u6e90",
            "updateNote": "\u6765\u81ea\u7528\u6237\u63d0\u4f9b\u7684\u5317\u4eac\u6587\u5316\u8d44\u6e90\u5bfc\u822a Excel\u3002\u7535\u8bdd\u3001\u5f00\u653e\u65f6\u95f4\u3001\u5165\u53e3\u94fe\u63a5\u548c\u5730\u5740\u4ecd\u5efa\u8bae\u4ee5\u5b98\u65b9\u9875\u9762\u518d\u6b21\u6838\u9a8c\u3002",
        }
        if resource_type in ["\u56fe\u4e66\u9986", "\u9605\u8bfb\u7a7a\u95f4"] or any(key in " ".join(tags) for key in ["\u9605\u8bfb", "\u81ea\u4e60"]):
            resource["reading"] = {
                "openTime": open_time or "\u4ee5\u5b98\u65b9\u4e3a\u51c6",
                "quietLevel": "\u5f85\u6838\u9a8c",
                "power": "\u5f85\u6838\u9a8c",
                "reservation": "\u4ee5\u5b98\u65b9\u4e3a\u51c6",
            }
        rows.append(resource)
    return rows


def build_events() -> list[dict]:
    wb = load_workbook(EVENT_XLSX, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if not raw or not text(raw[0]):
            continue
        (
            seq,
            date,
            time,
            weekday,
            period,
            is_weekend,
            venue,
            genre,
            category,
            title,
            performer,
            friendly_tag,
            map_url,
            show_url,
            ticket_url,
            note,
        ) = [text(item) for item in raw[:16]]
        tags = unique(split_tags(category) + split_tags(genre) + split_tags(friendly_tag) + ([period] if period else []))
        if is_weekend == "\u662f":
            tags.append("\u5468\u672b")
        joined = " ".join(tags + [title, genre, category])
        audiences = []
        if any(key in joined for key in ["\u513f\u7ae5", "\u4eb2\u5b50"]):
            audiences.append("\u4eb2\u5b50\u5bb6\u5ead")
        if any(key in joined for key in ["\u8f7b\u677e", "\u8131\u53e3\u79c0", "\u97f3\u4e50", "\u620f\u5267", "\u6c89\u6d78"]):
            audiences.append("\u4e0a\u73ed\u65cf")
        if any(key in joined for key in ["\u4f20\u7edf", "\u620f\u66f2", "\u76f8\u58f0", "\u4eac\u5267"]):
            audiences.append("\u8001\u5e74\u4eba")
        if any(key in joined for key in ["\u5b66\u751f", "\u4f4e\u6210\u672c"]):
            audiences.append("\u5b66\u751f")
        audiences.append("\u65b0\u5317\u4eac\u4eba/\u6e38\u5ba2")
        rows.append(
            {
                "id": f"event-{seq}",
                "title": title or "\u6f14\u51fa\u540d\u79f0\u5f85\u8865\u5145",
                "type": "\u6f14\u51fa",
                "date": date,
                "time": time or "\u4ee5\u5b98\u65b9\u4e3a\u51c6",
                "weekday": weekday,
                "period": period,
                "venue": venue or "\u573a\u9986\u5f85\u8865\u5145",
                "district": "\u5f85\u8865\u5145",
                "address": "\u5f85\u8865\u5145",
                "genre": genre or category or "\u5176\u4ed6",
                "performer": performer or "\u5f85\u8865\u5145",
                "price": "\u4ee5\u5b98\u65b9\u4e3a\u51c6",
                "tags": unique(tags),
                "audiences": unique(audiences),
                "ticketUrl": ticket_url or show_url,
                "mapUrl": map_url or baidu(venue),
                "amapUrl": amap(venue),
                "source": "\u5317\u4eac\u5e02\u6587\u5316\u548c\u65c5\u6e38\u5c40\u300a2026\u5e745\u67084\u65e5\u2014\u20145\u670810\u65e5\u5317\u4eac\u5546\u4e1a\u6f14\u51fa\u4fe1\u606f\u300bPDF\u6574\u7406\u6570\u636e",
                "verifyStatus": "\u9700\u6838\u9a8c",
                "note": note or "\u6f14\u51fa\u4fe1\u606f\u53ef\u80fd\u53d1\u751f\u53d8\u5316\uff0c\u4ee5\u5b98\u65b9\u9875\u9762\u6216\u73b0\u573a\u5b9e\u9645\u60c5\u51b5\u4e3a\u51c6\u3002",
            }
        )
    return rows


def write_files(resources: list[dict], events: list[dict]) -> None:
    districts = sorted({item["district"] for item in resources if item["district"] != "\u5f85\u8865\u5145"})
    genres = sorted({item["genre"] for item in events if item["genre"]})
    resources_ts = f'''export type VerifyStatus = "\\u5b98\\u65b9\\u6765\\u6e90" | "\\u5df2\\u6838\\u9a8c" | "\\u9700\\u4eba\\u5de5\\u6838\\u9a8c" | "Mock\\u793a\\u4f8b";

export type ResourceType =
  | "\\u56fe\\u4e66\\u9986"
  | "\\u6587\\u5316\\u9986"
  | "\\u535a\\u7269\\u9986/\\u7f8e\\u672f\\u9986"
  | "\\u5267\\u573a/\\u6f14\\u51fa"
  | "\\u516c\\u56ed/\\u666f\\u533a"
  | "\\u975e\\u9057\\u7a7a\\u95f4"
  | "\\u9605\\u8bfb\\u7a7a\\u95f4"
  | "\\u7535\\u5f71/\\u653e\\u6620"
  | "\\u5176\\u4ed6\\u6587\\u5316\\u7a7a\\u95f4";

export type AudienceType = "\\u5b66\\u751f" | "\\u4e0a\\u73ed\\u65cf" | "\\u8001\\u5e74\\u4eba" | "\\u4eb2\\u5b50\\u5bb6\\u5ead" | "\\u6b8b\\u969c\\u4eba\\u58eb" | "\\u65b0\\u5317\\u4eac\\u4eba/\\u6e38\\u5ba2";

export type Resource = {{
  id: string;
  name: string;
  type: ResourceType;
  district: string;
  address?: string;
  phone?: string;
  openTime?: string;
  lat?: number;
  lng?: number;
  tags: string[];
  audiences: AudienceType[];
  intro: string;
  image?: string;
  officialUrl?: string;
  activityUrl?: string;
  ticketUrl?: string;
  mapUrl: string;
  amapUrl: string;
  isFree?: boolean;
  accessibility?: {{ entrance?: boolean; toilet?: boolean; wheelchair?: boolean; lowServiceDesk?: boolean; brailleGuide?: boolean; accessibleRoute?: boolean; }};
  reading?: {{ openTime?: string; quietLevel?: string; power?: string; reservation?: string; }};
  source: string;
  verifyStatus: VerifyStatus;
  updateNote?: string;
}};

export const resourceTypes: ResourceType[] = ["\\u56fe\\u4e66\\u9986", "\\u6587\\u5316\\u9986", "\\u535a\\u7269\\u9986/\\u7f8e\\u672f\\u9986", "\\u5267\\u573a/\\u6f14\\u51fa", "\\u516c\\u56ed/\\u666f\\u533a", "\\u975e\\u9057\\u7a7a\\u95f4", "\\u9605\\u8bfb\\u7a7a\\u95f4", "\\u7535\\u5f71/\\u653e\\u6620", "\\u5176\\u4ed6\\u6587\\u5316\\u7a7a\\u95f4"];
export const districts = {dump(districts)};
export const audienceTypes: AudienceType[] = ["\\u5b66\\u751f", "\\u4e0a\\u73ed\\u65cf", "\\u8001\\u5e74\\u4eba", "\\u4eb2\\u5b50\\u5bb6\\u5ead", "\\u6b8b\\u969c\\u4eba\\u58eb", "\\u65b0\\u5317\\u4eac\\u4eba/\\u6e38\\u5ba2"];

export const resources: Resource[] = {dump(resources)};
'''
    events_ts = f'''export type EventItem = {{
  id: string;
  title: string;
  type: "\\u6f14\\u51fa" | "\\u5c55\\u89c8" | "\\u9605\\u8bfb\\u6d3b\\u52a8" | "\\u4eb2\\u5b50\\u6d3b\\u52a8";
  date: string;
  time: string;
  weekday?: string;
  period?: string;
  venue: string;
  district?: string;
  address?: string;
  genre: string;
  performer?: string;
  price?: string;
  tags: string[];
  audiences: string[];
  ticketUrl?: string;
  mapUrl: string;
  amapUrl: string;
  source: string;
  verifyStatus: "\\u5b98\\u65b9\\u4fe1\\u606f" | "\\u9700\\u6838\\u9a8c" | "Mock\\u793a\\u4f8b";
  note?: string;
}};

export const eventTypes = ["\\u6f14\\u51fa", "\\u5c55\\u89c8", "\\u9605\\u8bfb\\u6d3b\\u52a8", "\\u4eb2\\u5b50\\u6d3b\\u52a8"];
export const eventGenres = {dump(genres)};

export const events: EventItem[] = {dump(events)};
'''
    (ROOT / "src/data/resources.ts").write_text(resources_ts, encoding="utf-8")
    (ROOT / "src/data/events.ts").write_text(events_ts, encoding="utf-8")


if __name__ == "__main__":
    resources = build_resources()
    events = build_events()
    write_files(resources, events)
    print(f"resources={len(resources)} events={len(events)}")
